"""
Funções de visualização — Projetos SAF no Brasil.
Cada função neste arquivo:
  - Não recebe argumentos (para compatibilidade com o sistema de catálogo)
  - Retorna um objeto de visualização (plotly.graph_objects.Figure ou folium.Map)
  - Carrega os dados diretamente dos arquivos em dados/
Para adicionar uma nova visualização:
  1. Crie a função abaixo seguindo o padrão existente
  2. Registre-a no notebook: from src.registro import registrar; registrar(...)
"""

import warnings
import re
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import openpyxl
import geopandas as gpd
import folium
import plotly.graph_objects as go
from plotly.subplots import make_subplots

warnings.filterwarnings("ignore")

_PROJETO = Path(__file__).parent.parent
EXCEL = _PROJETO / "SAF_EPE_10projetos_Validado_atualizado5.xlsx"
SHP   = _PROJETO / "data" / "estados.geojson"


# ============================================================
# HELPERS COMPARTILHADOS
# ============================================================

def _fix_coord(val, kind):
    """Corrige coordenadas com escala errada dividindo por potências de 10."""
    if val is None:
        return None
    try:
        val = float(val)
    except (ValueError, TypeError):
        return None
    lo, hi = (-35.0, 6.0) if kind == "lat" else (-75.0, -28.0)
    if lo <= val <= hi:
        return val
    for p in range(1, 20):
        v = val / (10 ** p)
        if lo <= v <= hi:
            return v
    return None


def _carregar_df():
    """Lê o Excel via openpyxl e retorna DataFrame limpo com lat/lon corrigidos."""
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    df_raw = pd.DataFrame(list(ws.iter_rows(min_row=2, values_only=True)), columns=headers)

    col_map = {}
    for c in df_raw.columns:
        s = str(c)
        if "Projeto" in s and "EPE" in s:                col_map[c] = "Projeto"
        elif "Munic" in s:                               col_map[c] = "Municipio"
        elif "Capacidade" in s:                          col_map[c] = "Capacidade"
        elif "Rota" in s:                                col_map[c] = "Rota"
        elif "Feedstock" in s:                           col_map[c] = "Feedstock"
        elif "Ano" in s:                                 col_map[c] = "Ano"
        elif "Est" in s and "gio" in s and "Base" not in s: col_map[c] = "Estagio"
        elif "Base" in s and "Est" in s:                 col_map[c] = "BaseEstagio"

    df = df_raw.rename(columns=col_map).copy()
    df["lat"] = df["lat"].apply(lambda v: _fix_coord(v, "lat"))
    df["lon"] = df["lon"].apply(lambda v: _fix_coord(v, "lon"))
    df = df.dropna(subset=["lat", "lon"]).reset_index(drop=True)

    seen = {}
    for i, row in df.iterrows():
        key = (round(row["lat"], 2), round(row["lon"], 2))
        n = seen.get(key, 0)
        if n > 0:
            df.at[i, "lat"] += n * 0.09
            df.at[i, "lon"] += n * 0.09
        seen[key] = n + 1

    return df


def _parse_cap(v):
    s = str(v).replace("~", "").strip()
    m = re.match(r"^[\d\.,]+", s)
    if not m:
        return None
    try:
        return float(m.group(0).replace(".", "").replace(",", "."))
    except Exception:
        return None


def _fmt(v):
    """Formata número com ponto como separador de milhar (PT-BR)."""
    return f"{v:,.0f}".replace(",", ".")


# ============================================================
# MAPA FOLIUM — PORTUGUÊS
# ============================================================

def criar_mapa_saf_folium() -> folium.Map:
    """
    Mapa interativo Folium com os projetos SAF no Brasil.
    Marcadores numerados e coloridos por rota tecnológica, com popup de card detalhado.

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
           dados/BR_UF_2023/BR_UF_2023.shp
    Returns: folium.Map
    """
    df = _carregar_df()
    gdf_uf = gpd.read_file(SHP)

    PALETA_ROTA = {
        "Coprocessamento": "#1B4F8A",
        "HEFA":            "#2980B9",
        "ATJ":             "#E67E22",
        "FT":              "#1A7F4B",
    }

    def cor_rota(rota):
        for k, v in PALETA_ROTA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def card_html(row):
        projeto      = row.get("Projeto", "—")
        proponente   = row.get("Proponente", "—")
        capacidade   = row.get("Capacidade", "—")
        feedstock    = row.get("Feedstock", "—")
        rota         = row.get("Rota", "—")
        ano          = row.get("Ano", "—")
        investimento = row.get("Investimento", "—")
        base         = str(row.get("BaseEstagio", "—") or "—")
        if len(base) > 130:
            base = base[:127] + "..."
        url1  = str(row.get("URL Fonte 1", "") or "")
        url2  = str(row.get("URL Fonte 2", "") or "")
        nome1 = str(row.get("Fonte 1 (oficial)", "Fonte 1") or "Fonte 1")[:55]
        nome2 = str(row.get("Fonte 2", "Fonte 2") or "Fonte 2")[:55]

        fontes = ""
        ref = 1
        if url1.startswith("http"):
            fontes += f'<div style="margin-bottom:4px;">[{ref}] <a href="{url1}" target="_blank" style="color:#2C7BE5;text-decoration:none;">{nome1} ↗</a></div>'
            ref += 1
        if url2.startswith("http"):
            fontes += f'<div>[{ref}] <a href="{url2}" target="_blank" style="color:#2C7BE5;text-decoration:none;">{nome2} ↗</a></div>'

        def campo(label, valor):
            return (
                f'<div style="margin-bottom:9px;">'
                f'<div style="color:#6B7A8D;font-size:10px;text-transform:uppercase;'
                f'letter-spacing:.6px;font-weight:600;margin-bottom:2px;">{label}</div>'
                f'<div style="color:#1A2A3A;font-size:12px;font-weight:500;">{valor}</div>'
                f'</div>'
            )

        fontes_bloco = (
            f'<div style="padding:0 16px 12px 16px;border-top:1px solid #EEF2F7;'
            f'margin-top:4px;padding-top:10px;">'
            f'<div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;'
            f'letter-spacing:.6px;font-weight:700;margin-bottom:6px;">Fontes</div>'
            f'{fontes}</div>'
        ) if fontes else ""

        return f"""
<div style="font-family:'Segoe UI',Arial,sans-serif;width:360px;border-radius:10px;
     overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.18);border:1px solid #D8E3EC;background:white;">
  <div style="background:linear-gradient(135deg,#0D2E57 0%,#1B4F8A 100%);padding:14px 16px 12px;">
    <div style="font-size:9px;color:#7EB3E8;text-transform:uppercase;letter-spacing:1.2px;
         font-weight:700;margin-bottom:6px;">Projeto SAF — Brasil</div>
    <div style="font-size:16px;font-weight:700;color:#FFF;margin-bottom:3px;">{projeto}</div>
    <div style="font-size:12px;color:#A8C8EC;">{proponente}</div>
  </div>
  <div style="display:flex;background:#F0F5FB;border-bottom:1px solid #D8E3EC;">
    <div style="flex:1;padding:10px 16px;border-right:1px solid #D8E3EC;">
      <div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;letter-spacing:.6px;
           font-weight:700;margin-bottom:3px;">Capacidade SAF</div>
      <div style="color:#0D2E57;font-size:14px;font-weight:700;">{capacidade}</div>
      <div style="color:#6B7A8D;font-size:9px;">m³/ano</div>
    </div>
    <div style="flex:1;padding:10px 16px;">
      <div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;letter-spacing:.6px;
           font-weight:700;margin-bottom:3px;">Ano Início</div>
      <div style="color:#0D2E57;font-size:14px;font-weight:700;">{ano}</div>
    </div>
  </div>
  <div style="padding:12px 16px 4px 16px;">
    {campo("Feedstock Principal", feedstock)}
    {campo("Rota Tecnológica", rota)}
    {campo("Investimento", investimento)}
    {campo("Base do Estágio", base)}
  </div>
  {fontes_bloco}
</div>"""

    mapa = folium.Map(location=[-15.0, -52.0], zoom_start=4, tiles="CartoDB positron", control_scale=True)

    folium.GeoJson(
        gdf_uf.__geo_interface__,
        name="Estados",
        style_function=lambda _: {"fillColor": "transparent", "color": "#4A6FA5", "weight": 0.9, "fillOpacity": 0},
        tooltip=folium.GeoJsonTooltip(
            fields=["SIGLA_UF", "NM_UF"], aliases=["UF:", "Estado:"],
            style="font-family:sans-serif;font-size:12px;"
        ),
    ).add_to(mapa)

    grupo = folium.FeatureGroup(name="Projetos SAF", show=True)
    for i, (_, row) in enumerate(df.iterrows()):
        cor = cor_rota(row.get("Rota", ""))
        icon_html = (
            f'<div style="background:{cor};width:26px;height:26px;border-radius:50%;'
            f'border:3px solid white;box-shadow:0 2px 6px rgba(0,0,0,.35);'
            f'display:flex;align-items:center;justify-content:center;'
            f'font-size:9px;font-weight:bold;color:white;font-family:sans-serif;">'
            f'{i + 1}</div>'
        )
        iframe = folium.IFrame(html=card_html(row), width=390, height=520)
        grupo.add_child(folium.Marker(
            location=[row["lat"], row["lon"]],
            popup=folium.Popup(iframe, max_width=400),
            tooltip=folium.Tooltip(
                f'<b>{row.get("Projeto", "")}</b>',
                style="font-family:sans-serif;font-size:13px;"
            ),
            icon=folium.DivIcon(html=icon_html, icon_size=(32, 32), icon_anchor=(16, 16)),
        ))

    mapa.add_child(grupo)
    folium.LayerControl(collapsed=False).add_to(mapa)

    mapa.get_root().html.add_child(folium.Element("""
<div style="position:fixed;bottom:30px;left:30px;background:white;padding:12px 16px;
     border-radius:8px;border:1px solid #D8E3EC;font-family:'Segoe UI',sans-serif;
     font-size:12px;box-shadow:0 2px 8px rgba(0,0,0,.15);z-index:1000;">
  <div style="font-weight:700;margin-bottom:8px;color:#1A2A3A;">Rota Tecnológica</div>
  <div><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#1B4F8A;margin-right:6px;vertical-align:middle;"></span>Coprocessamento HEFA</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#2980B9;margin-right:6px;vertical-align:middle;"></span>HEFA Dedicado</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#E67E22;margin-right:6px;vertical-align:middle;"></span>ATJ (Alcohol-to-Jet)</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#607D8B;margin-right:6px;vertical-align:middle;"></span>Outros</div>
</div>"""))

    return mapa


# ============================================================
# MAPA FOLIUM — ENGLISH
# ============================================================

def create_saf_map_folium() -> folium.Map:
    """
    Interactive Folium map of SAF projects in Brazil.
    Numbered markers colored by technology route, with detailed card popups.

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
            dados/BR_UF_2023/BR_UF_2023.shp
    Returns: folium.Map
    """
    df = _carregar_df()
    gdf_uf = gpd.read_file(SHP)

    PALETA_ROTA = {
        "Coprocessamento": "#1B4F8A",
        "HEFA":            "#2980B9",
        "ATJ":             "#E67E22",
        "FT":              "#1A7F4B",
    }

    def cor_rota(rota):
        for k, v in PALETA_ROTA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def card_html(row):
        project      = row.get("Projeto", "—")
        proponent    = row.get("Proponente", "—")
        capacity     = row.get("Capacidade", "—")
        feedstock    = row.get("Feedstock", "—")
        route        = row.get("Rota", "—")
        year         = row.get("Ano", "—")
        investment   = row.get("Investimento", "—")
        base         = str(row.get("BaseEstagio", "—") or "—")
        if len(base) > 130:
            base = base[:127] + "..."
        url1  = str(row.get("URL Fonte 1", "") or "")
        url2  = str(row.get("URL Fonte 2", "") or "")
        name1 = str(row.get("Fonte 1 (oficial)", "Source 1") or "Source 1")[:55]
        name2 = str(row.get("Fonte 2", "Source 2") or "Source 2")[:55]

        sources = ""
        ref = 1
        if url1.startswith("http"):
            sources += f'<div style="margin-bottom:4px;">[{ref}] <a href="{url1}" target="_blank" style="color:#2C7BE5;text-decoration:none;">{name1} ↗</a></div>'
            ref += 1
        if url2.startswith("http"):
            sources += f'<div>[{ref}] <a href="{url2}" target="_blank" style="color:#2C7BE5;text-decoration:none;">{name2} ↗</a></div>'

        def field(label, value):
            return (
                f'<div style="margin-bottom:9px;">'
                f'<div style="color:#6B7A8D;font-size:10px;text-transform:uppercase;'
                f'letter-spacing:.6px;font-weight:600;margin-bottom:2px;">{label}</div>'
                f'<div style="color:#1A2A3A;font-size:12px;font-weight:500;">{value}</div>'
                f'</div>'
            )

        sources_block = (
            f'<div style="padding:0 16px 12px 16px;border-top:1px solid #EEF2F7;'
            f'margin-top:4px;padding-top:10px;">'
            f'<div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;'
            f'letter-spacing:.6px;font-weight:700;margin-bottom:6px;">Sources</div>'
            f'{sources}</div>'
        ) if sources else ""

        return f"""
<div style="font-family:'Segoe UI',Arial,sans-serif;width:360px;border-radius:10px;
     overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.18);border:1px solid #D8E3EC;background:white;">
  <div style="background:linear-gradient(135deg,#0D2E57 0%,#1B4F8A 100%);padding:14px 16px 12px;">
    <div style="font-size:9px;color:#7EB3E8;text-transform:uppercase;letter-spacing:1.2px;
         font-weight:700;margin-bottom:6px;">SAF Project — Brazil</div>
    <div style="font-size:16px;font-weight:700;color:#FFF;margin-bottom:3px;">{project}</div>
    <div style="font-size:12px;color:#A8C8EC;">{proponent}</div>
  </div>
  <div style="display:flex;background:#F0F5FB;border-bottom:1px solid #D8E3EC;">
    <div style="flex:1;padding:10px 16px;border-right:1px solid #D8E3EC;">
      <div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;letter-spacing:.6px;
           font-weight:700;margin-bottom:3px;">SAF Capacity</div>
      <div style="color:#0D2E57;font-size:14px;font-weight:700;">{capacity}</div>
      <div style="color:#6B7A8D;font-size:9px;">m³/year</div>
    </div>
    <div style="flex:1;padding:10px 16px;">
      <div style="color:#6B7A8D;font-size:9px;text-transform:uppercase;letter-spacing:.6px;
           font-weight:700;margin-bottom:3px;">Start Year</div>
      <div style="color:#0D2E57;font-size:14px;font-weight:700;">{year}</div>
    </div>
  </div>
  <div style="padding:12px 16px 4px 16px;">
    {field("Main Feedstock", feedstock)}
    {field("Technology Route", route)}
    {field("Investment", investment)}
    {field("Stage Basis", base)}
  </div>
  {sources_block}
</div>"""

    mapa = folium.Map(location=[-15.0, -52.0], zoom_start=4, tiles="CartoDB positron", control_scale=True)

    folium.GeoJson(
        gdf_uf.__geo_interface__,
        name="States",
        style_function=lambda _: {"fillColor": "transparent", "color": "#4A6FA5", "weight": 0.9, "fillOpacity": 0},
        tooltip=folium.GeoJsonTooltip(
            fields=["SIGLA_UF", "NM_UF"], aliases=["State:", "Name:"],
            style="font-family:sans-serif;font-size:12px;"
        ),
    ).add_to(mapa)

    grupo = folium.FeatureGroup(name="SAF Projects", show=True)
    for i, (_, row) in enumerate(df.iterrows()):
        cor = cor_rota(row.get("Rota", ""))
        icon_html = (
            f'<div style="background:{cor};width:26px;height:26px;border-radius:50%;'
            f'border:3px solid white;box-shadow:0 2px 6px rgba(0,0,0,.35);'
            f'display:flex;align-items:center;justify-content:center;'
            f'font-size:9px;font-weight:bold;color:white;font-family:sans-serif;">'
            f'{i + 1}</div>'
        )
        iframe = folium.IFrame(html=card_html(row), width=390, height=520)
        grupo.add_child(folium.Marker(
            location=[row["lat"], row["lon"]],
            popup=folium.Popup(iframe, max_width=400),
            tooltip=folium.Tooltip(
                f'<b>{row.get("Projeto", "")}</b>',
                style="font-family:sans-serif;font-size:13px;"
            ),
            icon=folium.DivIcon(html=icon_html, icon_size=(32, 32), icon_anchor=(16, 16)),
        ))

    mapa.add_child(grupo)
    folium.LayerControl(collapsed=False).add_to(mapa)

    mapa.get_root().html.add_child(folium.Element("""
<div style="position:fixed;bottom:30px;left:30px;background:white;padding:12px 16px;
     border-radius:8px;border:1px solid #D8E3EC;font-family:'Segoe UI',sans-serif;
     font-size:12px;box-shadow:0 2px 8px rgba(0,0,0,.15);z-index:1000;">
  <div style="font-weight:700;margin-bottom:8px;color:#1A2A3A;">Technology Route</div>
  <div><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#1B4F8A;margin-right:6px;vertical-align:middle;"></span>Co-processing HEFA</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#2980B9;margin-right:6px;vertical-align:middle;"></span>Dedicated HEFA</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#E67E22;margin-right:6px;vertical-align:middle;"></span>ATJ (Alcohol-to-Jet)</div>
  <div style="margin-top:4px;"><span style="display:inline-block;width:12px;height:12px;border-radius:50%;
       background:#607D8B;margin-right:6px;vertical-align:middle;"></span>Other</div>
</div>"""))

    return mapa


# ============================================================
# HELPERS PARA GRÁFICOS DE BARRAS + BOLHAS
# ============================================================

def _preparar_dados_graficos():
    """Prepara dados agregados para barras empilhadas e bolhas concêntricas."""
    df = _carregar_df()

    PALETA = {
        "Coprocessamento": "#b2c73c",
        "HEFA":            "#3357ff",
        "ATJ":             "#107c42",
        "FT":              "#1A7F4B",
    }

    def _cor(rota):
        for k, v in PALETA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def _norm_r(r):
        r = str(r)
        if "Copro" in r: return "Coprocessamento HEFA"
        if "ATJ"   in r: return "ATJ"
        if "HEFA"  in r: return "HEFA Dedicado"
        if "FT"    in r: return "FT"
        return "Outros"

    _df = df.copy()
    _df["_a"] = pd.to_numeric(_df["Ano"], errors="coerce")
    _df["_c"] = _df["Capacidade"].apply(_parse_cap)
    _df["_r"] = _df["Rota"].apply(_norm_r)
    _df = _df.dropna(subset=["_a", "_c"])
    _df["_a"] = _df["_a"].astype(int)
    _df = _df[_df["_a"].between(2025, 2037)]

    pvt   = (_df.groupby(["_a", "_r"])["_c"].sum()
               .unstack(fill_value=0)
               .reindex(sorted(_df["_a"].unique()), fill_value=0))
    pvt_k = pvt / 1_000
    tot_k = pvt_k.sum(axis=1)
    Y_MAX = tot_k.max()

    tr = _df.groupby("_r")["_c"].sum().sort_values(ascending=False)

    ORDEM = ["Coprocessamento HEFA", "HEFA Dedicado", "ATJ"]
    ROTAS = [r for r in ORDEM if r in pvt_k.columns]

    return pvt_k, tot_k, tr, ROTAS, Y_MAX, _cor


# ============================================================
# GRÁFICO BARRAS + BOLHAS — PORTUGUÊS
# ============================================================

def criar_grafico_barras_bolhas_saf() -> go.Figure:
    """
    Gráfico combinado — barras empilhadas por rota e ano de operação (esquerda)
    e bolhas concêntricas com capacidade total acumulada por rota (direita).

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    pvt_k, tot_k, tr, ROTAS, Y_MAX, _cor = _preparar_dados_graficos()
    ANOS_S = [str(a) for a in pvt_k.index]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Coprocessamento",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }
    CIRCULO_NOME = {
        "Coprocessamento HEFA": "Coproces.",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }
    CIRCULO_COR_TEXTO = {
        "Coprocessamento HEFA": "black",
        "HEFA Dedicado":        "white",
        "ATJ":                  "white",
    }

    fig = make_subplots(
        rows=1, cols=2,
        column_widths=[0.65, 0.35],
        horizontal_spacing=0.04,
        subplot_titles=[
            "Capacidade por Rota Tecnológica e Ano de Operação",
            "Capacidade Total por Rota (mil m³/ano)",
        ],
    )

    for rota in ROTAS:
        vals = pvt_k[rota].values
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=LEGENDA_LABEL.get(rota, rota),
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{rota}</b><br>"
                "Ano: %{x}<br>"
                "Capacidade: <b>%{y:,.0f} mil m³/ano</b>"
                "<extra></extra>"
            ),
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_k.values + Y_MAX * 0.045,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_k],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ), row=1, col=1)

    N       = len(tr)
    max_cap = tr.max()
    MAX_SZ  = 130
    MIN_SZ  = 70
    y_pos   = np.linspace(0.82, 0.18, N)
    PX      = 1 / 600

    for k, (rota, cap) in enumerate(tr.items()):
        cap_k   = cap / 1_000
        cor     = _cor(rota)
        sz      = MIN_SZ + (MAX_SZ - MIN_SZ) * np.sqrt(cap / max_cap)
        ly      = y_pos[k]
        nome    = CIRCULO_NOME.get(rota, rota)
        txt_cor = CIRCULO_COR_TEXTO.get(rota, "white")
        cap_cor = "rgba(0,0,0,0.75)" if txt_cor == "black" else "rgba(255,255,255,0.90)"
        offset  = 8 * PX

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly],
            mode="markers",
            showlegend=False,
            marker=dict(size=sz, color=cor, opacity=0.93, line=dict(color="white", width=3)),
            hovertemplate=(
                f"<b>{rota}</b><br>"
                f"Total: <b>{_fmt(cap_k)} mil m³/ano</b><extra></extra>"
            ),
        ), row=1, col=2)

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly + offset],
            mode="text",
            text=[f"<b>{nome}</b>"],
            textfont=dict(size=12, color=txt_cor, family="Arial, sans-serif"),
            textposition="middle center",
            showlegend=False, hoverinfo="skip",
        ), row=1, col=2)

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly - offset],
            mode="text",
            text=[f"{_fmt(cap_k)} mil m³/ano"],
            textfont=dict(size=10, color=cap_cor, family="Arial, sans-serif"),
            textposition="middle center",
            showlegend=False, hoverinfo="skip",
        ), row=1, col=2)

    fig.update_layout(
        barmode="stack",
        separators=",.",
        title=dict(
            text=(
                "<b>Projetos SAF no Brasil — Capacidade por Rota Tecnológica</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Capacidade instalada (mil m³ SAF/ano) · barras empilhadas · círculos = total acumulado por rota"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=600,
        margin=dict(l=80, r=20, t=95, b=130),
        legend=dict(
            title=dict(text="<b>Rota Tecnológica</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(0,0,0,0)",
            orientation="h",
            x=0.5, y=-0.25, xanchor="center",
        ),
        hoverlabel=dict(bgcolor="white", bordercolor="#D0D8E4", font=dict(size=12, family="Arial, sans-serif")),
        font=dict(family="Arial, sans-serif"),
    )

    fig.update_xaxes(
        tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
        title=dict(text="Ano de início de operação", font=dict(size=13, color="#555", family="Arial, sans-serif")),
        showgrid=False, zeroline=False, linecolor="#CCCCCC", linewidth=1,
        row=1, col=1,
    )
    fig.update_yaxes(
        title=dict(text="Capacidade (mil m³ SAF/ano)", font=dict(size=13, color="#555", family="Arial, sans-serif")),
        tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
        zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
        tickformat=",", rangemode="tozero", range=[0, Y_MAX * 1.20],
        row=1, col=1,
    )
    fig.update_xaxes(range=[0.05, 0.95], showgrid=False, zeroline=False, showticklabels=False, fixedrange=True, row=1, col=2)
    fig.update_yaxes(range=[0.0, 1.0], showgrid=False, zeroline=False, showticklabels=False, fixedrange=True, row=1, col=2)
    fig.update_annotations(font=dict(size=13, color="#0D2E57", family="Arial, sans-serif"))

    return fig


# ============================================================
# GRÁFICO BARRAS + BOLHAS — ENGLISH
# ============================================================

def create_bar_bubble_chart_saf() -> go.Figure:
    """
    Combined chart — stacked bars by technology route and operation year (left)
    and concentric bubbles with total accumulated capacity by route (right).

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    pvt_k, tot_k, tr, ROTAS, Y_MAX, _cor = _preparar_dados_graficos()
    ANOS_S = [str(a) for a in pvt_k.index]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Co-processing HEFA",
        "HEFA Dedicado":        "Dedicated HEFA",
        "ATJ":                  "ATJ",
    }
    CIRCULO_NOME = {
        "Coprocessamento HEFA": "Co-proc.",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }
    CIRCULO_COR_TEXTO = {
        "Coprocessamento HEFA": "black",
        "HEFA Dedicado":        "white",
        "ATJ":                  "white",
    }

    fig = make_subplots(
        rows=1, cols=2,
        column_widths=[0.65, 0.35],
        horizontal_spacing=0.04,
        subplot_titles=[
            "Capacity by Technology Route and Operation Year",
            "Total Capacity by Route (thousand m³/year)",
        ],
    )

    for rota in ROTAS:
        vals = pvt_k[rota].values
        label = LEGENDA_LABEL.get(rota, rota)
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=label,
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{label}</b><br>"
                "Year: %{x}<br>"
                "Capacity: <b>%{y:,.0f} thousand m³/year</b>"
                "<extra></extra>"
            ),
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_k.values + Y_MAX * 0.045,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_k],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ), row=1, col=1)

    N       = len(tr)
    max_cap = tr.max()
    MAX_SZ  = 130
    MIN_SZ  = 70
    y_pos   = np.linspace(0.82, 0.18, N)
    PX      = 1 / 600

    for k, (rota, cap) in enumerate(tr.items()):
        cap_k   = cap / 1_000
        cor     = _cor(rota)
        sz      = MIN_SZ + (MAX_SZ - MIN_SZ) * np.sqrt(cap / max_cap)
        ly      = y_pos[k]
        nome    = CIRCULO_NOME.get(rota, rota)
        txt_cor = CIRCULO_COR_TEXTO.get(rota, "white")
        cap_cor = "rgba(0,0,0,0.75)" if txt_cor == "black" else "rgba(255,255,255,0.90)"
        offset  = 8 * PX
        label   = LEGENDA_LABEL.get(rota, rota)

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly],
            mode="markers",
            showlegend=False,
            marker=dict(size=sz, color=cor, opacity=0.93, line=dict(color="white", width=3)),
            hovertemplate=(
                f"<b>{label}</b><br>"
                f"Total: <b>{_fmt(cap_k)} thousand m³/year</b><extra></extra>"
            ),
        ), row=1, col=2)

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly + offset],
            mode="text",
            text=[f"<b>{nome}</b>"],
            textfont=dict(size=12, color=txt_cor, family="Arial, sans-serif"),
            textposition="middle center",
            showlegend=False, hoverinfo="skip",
        ), row=1, col=2)

        fig.add_trace(go.Scatter(
            x=[0.50], y=[ly - offset],
            mode="text",
            text=[f"{_fmt(cap_k)} th. m³/yr"],
            textfont=dict(size=10, color=cap_cor, family="Arial, sans-serif"),
            textposition="middle center",
            showlegend=False, hoverinfo="skip",
        ), row=1, col=2)

    fig.update_layout(
        barmode="stack",
        title=dict(
            text=(
                "<b>SAF Projects in Brazil — Capacity by Technology Route</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Installed capacity (thousand m³ SAF/year) · stacked bars · bubbles = total accumulated capacity by route"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=600,
        margin=dict(l=80, r=20, t=95, b=130),
        legend=dict(
            title=dict(text="<b>Technology Route</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(0,0,0,0)",
            orientation="h",
            x=0.5, y=-0.25, xanchor="center",
        ),
        hoverlabel=dict(bgcolor="white", bordercolor="#D0D8E4", font=dict(size=12, family="Arial, sans-serif")),
        font=dict(family="Arial, sans-serif"),
    )

    fig.update_xaxes(
        tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
        title=dict(text="Operation start year", font=dict(size=13, color="#555", family="Arial, sans-serif")),
        showgrid=False, zeroline=False, linecolor="#CCCCCC", linewidth=1,
        row=1, col=1,
    )
    fig.update_yaxes(
        title=dict(text="Capacity (thousand m³ SAF/year)", font=dict(size=13, color="#555", family="Arial, sans-serif")),
        tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
        zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
        tickformat=",", rangemode="tozero", range=[0, Y_MAX * 1.20],
        row=1, col=1,
    )
    fig.update_xaxes(range=[0.05, 0.95], showgrid=False, zeroline=False, showticklabels=False, fixedrange=True, row=1, col=2)
    fig.update_yaxes(range=[0.0, 1.0], showgrid=False, zeroline=False, showticklabels=False, fixedrange=True, row=1, col=2)
    fig.update_annotations(font=dict(size=13, color="#0D2E57", family="Arial, sans-serif"))

    return fig


# ============================================================
# GRÁFICO APENAS BARRAS EMPILHADAS — PORTUGUÊS
# ============================================================

def criar_grafico_barras_saf() -> go.Figure:
    """
    Barras empilhadas: capacidade SAF por rota tecnológica e ano de início de operação.

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    pvt_k, tot_k, tr, ROTAS, Y_MAX, _cor = _preparar_dados_graficos()
    ANOS_S = [str(a) for a in pvt_k.index]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Coprocessamento",
        "HEFA Dedicado":        "HEFA Dedicado",
        "ATJ":                  "ATJ",
    }

    fig = go.Figure()

    for rota in ROTAS:
        vals = pvt_k[rota].values
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=LEGENDA_LABEL.get(rota, rota),
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{rota}</b><br>"
                "Ano: %{x}<br>"
                "Capacidade: <b>%{y:,.0f} mil m³/ano</b>"
                "<extra></extra>"
            ),
        ))

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_k.values + Y_MAX * 0.045,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_k],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ))

    fig.update_layout(
        barmode="stack",
        separators=",.",
        title=dict(
            text=(
                "<b>Projetos SAF no Brasil — Capacidade por Rota e Ano de Operação</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Capacidade instalada (mil m³ SAF/ano) · barras empilhadas por rota tecnológica"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=560,
        margin=dict(l=80, r=30, t=95, b=120),
        xaxis=dict(
            title=dict(text="Ano de início de operação", font=dict(size=13, color="#555", family="Arial, sans-serif")),
            tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
            showgrid=False, zeroline=False, linecolor="#CCCCCC", linewidth=1,
        ),
        yaxis=dict(
            title=dict(text="Capacidade (mil m³ SAF/ano)", font=dict(size=13, color="#555", family="Arial, sans-serif")),
            tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
            showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
            zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
            tickformat=",", rangemode="tozero", range=[0, Y_MAX * 1.20],
        ),
        legend=dict(
            title=dict(text="<b>Rota Tecnológica</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            orientation="h",
            x=0.5, y=-0.22, xanchor="center",
        ),
        hoverlabel=dict(bgcolor="white", bordercolor="#D0D8E4", font=dict(size=12, family="Arial, sans-serif")),
        font=dict(family="Arial, sans-serif"),
    )

    return fig


# ============================================================
# GRÁFICO APENAS BARRAS EMPILHADAS — ENGLISH
# ============================================================

def create_stacked_bar_chart_saf() -> go.Figure:
    """
    Stacked bar chart: SAF capacity by technology route and operation start year.

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    pvt_k, tot_k, tr, ROTAS, Y_MAX, _cor = _preparar_dados_graficos()
    ANOS_S = [str(a) for a in pvt_k.index]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Co-processing HEFA",
        "HEFA Dedicado":        "Dedicated HEFA",
        "ATJ":                  "ATJ",
    }

    fig = go.Figure()

    for rota in ROTAS:
        vals = pvt_k[rota].values
        label = LEGENDA_LABEL.get(rota, rota)
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=label,
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{label}</b><br>"
                "Year: %{x}<br>"
                "Capacity: <b>%{y:,.0f} thousand m³/year</b>"
                "<extra></extra>"
            ),
        ))

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_k.values + Y_MAX * 0.045,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_k],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ))

    fig.update_layout(
        barmode="stack",
        title=dict(
            text=(
                "<b>SAF Projects in Brazil — Capacity by Route and Operation Year</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Installed capacity (thousand m³ SAF/year) · stacked bars by technology route"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=560,
        margin=dict(l=80, r=30, t=95, b=120),
        xaxis=dict(
            title=dict(text="Operation start year", font=dict(size=13, color="#555", family="Arial, sans-serif")),
            tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
            showgrid=False, zeroline=False, linecolor="#CCCCCC", linewidth=1,
        ),
        yaxis=dict(
            title=dict(text="Capacity (thousand m³ SAF/year)", font=dict(size=13, color="#555", family="Arial, sans-serif")),
            tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
            showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
            zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
            tickformat=",", rangemode="tozero", range=[0, Y_MAX * 1.20],
        ),
        legend=dict(
            title=dict(text="<b>Technology Route</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            orientation="h",
            x=0.5, y=-0.22, xanchor="center",
        ),
        hoverlabel=dict(bgcolor="white", bordercolor="#D0D8E4", font=dict(size=12, family="Arial, sans-serif")),
        font=dict(family="Arial, sans-serif"),
    )

    return fig


# ============================================================
# GRÁFICO LINHA DO TEMPO — PORTUGUÊS
# ============================================================

def criar_grafico_timeline_saf() -> go.Figure:
    """
    Linha do tempo dos projetos SAF no Brasil: cards posicionados por ano de início,
    com hover contendo detalhes do projeto.

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    df = _carregar_df()

    def _cor_tl(rota):
        r = str(rota)
        if "Copro" in r: return "#1B4F8A"
        if "ATJ"   in r: return "#E67E22"
        if "HEFA"  in r: return "#2980B9"
        return "#78909C"

    def _abrev(texto, n=20):
        s = (str(texto)
             .replace("Petrobras ", "PBR ")
             .replace(" (Copro)", "")
             .replace(" (ATJ)", ""))
        return (s[:n - 1] + "…") if len(s) > n else s

    df_tl = df.copy()
    df_tl["_ano"] = pd.to_numeric(df_tl["Ano"], errors="coerce")
    df_com = df_tl[df_tl["_ano"].notna()].sort_values("_ano").reset_index(drop=True)
    df_sem = df_tl[df_tl["_ano"].isna()].reset_index(drop=True)

    LEVELS = [2.6, -2.6, 5.4, -5.4]
    _cnt   = defaultdict(int)
    card_pos = []
    for _, row in df_com.iterrows():
        ano = int(row["_ano"])
        k   = _cnt[ano]
        card_pos.append((ano, LEVELS[k % len(LEVELS)]))
        _cnt[ano] += 1

    CW, CH, CHH = 0.84, 1.85, 0.58
    X0_TL, X1_TL   = 2024.3, 2033.8
    X0_SEM, X1_SEM = 2034.6, 2037.6

    shapes, annotations = [], []
    hx, hy, ht = [], [], []

    shapes.append(dict(
        type="line", x0=X0_TL, x1=X1_TL, y0=0, y1=0,
        line=dict(color="#0D2E57", width=4),
    ))

    for ano in sorted(df_com["_ano"].unique()):
        a = int(ano)
        shapes.append(dict(
            type="circle",
            x0=a - 0.17, x1=a + 0.17, y0=-0.17, y1=0.17,
            fillcolor="#0D2E57", line=dict(color="white", width=2),
            layer="above",
        ))
        annotations.append(dict(
            x=a, y=-0.60, text=f"<b>{a}</b>",
            showarrow=False, xanchor="center", yanchor="top",
            font=dict(size=12, color="#0D2E57", family="Arial, sans-serif"),
        ))

    for i, (_, row) in enumerate(df_com.iterrows()):
        cx, cy = card_pos[i]
        cor    = _cor_tl(row.get("Rota", ""))
        x0, x1 = cx - CW / 2, cx + CW / 2
        y0, y1 = cy - CH / 2, cy + CH / 2
        yh0    = y1 - CHH

        y_link = y0 if cy > 0 else y1
        shapes.append(dict(
            type="line",
            x0=cx, x1=cx,
            y0=(0.19 if cy > 0 else -0.19), y1=y_link,
            line=dict(color=cor, width=1.5, dash="dot"),
        ))
        shapes.append(dict(
            type="rect", x0=x0, x1=x1, y0=y0, y1=yh0,
            fillcolor="white", line=dict(color="#C5D5E5", width=0.8),
            layer="above",
        ))
        shapes.append(dict(
            type="rect", x0=x0, x1=x1, y0=yh0, y1=y1,
            fillcolor=cor, line=dict(color="rgba(255,255,255,.4)", width=0.5),
            layer="above",
        ))
        annotations += [
            dict(x=cx, y=y1 - 0.09, text=f'<b>{_abrev(row.get("Projeto", ""))}</b>',
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=8, color="white", family="Arial, sans-serif")),
            dict(x=cx, y=yh0 - 0.10, text=str(row.get("Capacidade", "—")) + " m³/ano",
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=7.5, color="#1A2A3A", family="Arial, sans-serif")),
            dict(x=cx, y=yh0 - 0.52, text=f'<i>{_abrev(row.get("Rota", ""), 22)}</i>',
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=6.5, color="#555", family="Arial, sans-serif")),
        ]
        hx.append(cx); hy.append(cy)
        ht.append(
            f'<b>{row.get("Projeto", "")}</b><br>'
            f'Proponente: {row.get("Proponente", "")}<br>'
            f'Capacidade: {row.get("Capacidade", "")} m³/ano<br>'
            f'Rota: {row.get("Rota", "")}<br>'
            f'Feedstock: {row.get("Feedstock", "")}<br>'
            f'Início previsto: {row.get("Ano", "")}<br>'
            f'Estágio: {row.get("Estagio", "")}<br>'
            f'Investimento: {row.get("Investimento", "")}'
        )

    if len(df_sem) > 0:
        shapes.append(dict(
            type="line",
            x0=X0_SEM - 0.35, x1=X0_SEM - 0.35, y0=-7.8, y1=7.8,
            line=dict(color="#BBBBBB", width=1.5, dash="longdash"),
        ))
        annotations.append(dict(
            x=(X0_SEM + X1_SEM) / 2, y=7.5,
            text="<b>Sem previsão<br>definida</b>",
            showarrow=False, xanchor="center", yanchor="top", align="center",
            font=dict(size=11, color="#888", family="Arial, sans-serif"),
        ))
        cx_sem = (X0_SEM + X1_SEM) / 2
        for j, (_, row) in enumerate(df_sem.iterrows()):
            cy_sem = 5.5 - j * (CH + 0.40)
            cor    = _cor_tl(row.get("Rota", ""))
            xs0, xs1 = cx_sem - CW / 2, cx_sem + CW / 2
            ys0, ys1 = cy_sem - CH / 2, cy_sem + CH / 2
            yhs0     = ys1 - CHH

            shapes += [
                dict(type="rect", x0=xs0, x1=xs1, y0=yhs0, y1=ys1,
                     fillcolor=cor, line=dict(color="rgba(255,255,255,.4)", width=0.5),
                     layer="above"),
                dict(type="rect", x0=xs0, x1=xs1, y0=ys0, y1=yhs0,
                     fillcolor="#F5F5F5", line=dict(color="#C5D5E5", width=0.8),
                     layer="above"),
            ]
            annotations += [
                dict(x=cx_sem, y=ys1 - 0.09, text=f'<b>{_abrev(row.get("Projeto", ""))}</b>',
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=8, color="white", family="Arial, sans-serif")),
                dict(x=cx_sem, y=yhs0 - 0.10, text=str(row.get("Capacidade", "—")) + " m³/ano",
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=7.5, color="#1A2A3A", family="Arial, sans-serif")),
                dict(x=cx_sem, y=yhs0 - 0.52, text=f'<i>{_abrev(row.get("Rota", ""), 22)}</i>',
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=6.5, color="#555", family="Arial, sans-serif")),
            ]
            hx.append(cx_sem); hy.append(cy_sem)
            ht.append(
                f'<b>{row.get("Projeto", "")}</b><br>'
                f'Proponente: {row.get("Proponente", "")}<br>'
                f'Capacidade: {row.get("Capacidade", "")} m³/ano<br>'
                f'Rota: {row.get("Rota", "")}<br>'
                f'Feedstock: {row.get("Feedstock", "")}<br>'
                f'Início previsto: <b>Sem previsão definida</b><br>'
                f'Estágio: {row.get("Estagio", "")}<br>'
                f'Investimento: {row.get("Investimento", "")}'
            )

    _legenda = [
        ("Coprocessamento HEFA", "#1B4F8A"),
        ("HEFA Dedicado",        "#2980B9"),
        ("ATJ (Alcohol-to-Jet)", "#E67E22"),
        ("Outros",               "#78909C"),
    ]
    annotations.append(dict(
        x=X0_TL + 0.05, y=-5.8, text="<b>Rota Tecnológica</b>",
        showarrow=False, xanchor="left", yanchor="top",
        font=dict(size=9.5, color="#0D2E57", family="Arial, sans-serif"),
    ))
    for k, (nome_r, cor_r) in enumerate(_legenda):
        lx = X0_TL + 0.05
        ly = -6.4 - k * 0.72
        shapes.append(dict(
            type="circle",
            x0=lx - 0.12, x1=lx + 0.12, y0=ly - 0.12, y1=ly + 0.12,
            fillcolor=cor_r, line=dict(color="white", width=1),
            layer="above",
        ))
        annotations.append(dict(
            x=lx + 0.22, y=ly, text=nome_r,
            showarrow=False, xanchor="left", yanchor="middle",
            font=dict(size=9, color="#333", family="Arial, sans-serif"),
        ))

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=hx, y=hy, mode="markers",
        marker=dict(size=44, opacity=0, symbol="square"),
        text=ht,
        hovertemplate="%{text}<extra></extra>",
        showlegend=False,
    ))
    fig.update_layout(
        shapes=shapes,
        annotations=annotations,
        title=dict(
            text=(
                "<b>Linha do Tempo — Projetos SAF no Brasil</b><br>"
                '<span style="font-size:11px;color:#666">'
                "Ano previsto de início de operação · Passe o mouse sobre os cards para detalhes"
                "</span>"
            ),
            font=dict(size=15, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        xaxis=dict(
            range=[X0_TL - 0.5, X1_SEM + 0.3],
            showgrid=False, zeroline=False, showticklabels=False, fixedrange=True,
        ),
        yaxis=dict(
            range=[-8.2, 8.2],
            showgrid=False, zeroline=False, showticklabels=False, fixedrange=True,
        ),
        paper_bgcolor="#F4F7FB",
        plot_bgcolor="#F4F7FB",
        height=700,
        margin=dict(l=10, r=10, t=90, b=10),
        showlegend=False,
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=11, family="Arial, sans-serif"),
        ),
    )
    return fig


# ============================================================
# GRÁFICO LINHA DO TEMPO — ENGLISH
# ============================================================

def create_timeline_chart_saf() -> go.Figure:
    """
    Timeline of SAF projects in Brazil: cards positioned by operation start year,
    with hover details for each project.

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    df = _carregar_df()

    def _cor_tl(rota):
        r = str(rota)
        if "Copro" in r: return "#1B4F8A"
        if "ATJ"   in r: return "#E67E22"
        if "HEFA"  in r: return "#2980B9"
        return "#78909C"

    def _abrev(texto, n=20):
        s = (str(texto)
             .replace("Petrobras ", "PBR ")
             .replace(" (Copro)", "")
             .replace(" (ATJ)", ""))
        return (s[:n - 1] + "…") if len(s) > n else s

    df_tl = df.copy()
    df_tl["_ano"] = pd.to_numeric(df_tl["Ano"], errors="coerce")
    df_com = df_tl[df_tl["_ano"].notna()].sort_values("_ano").reset_index(drop=True)
    df_sem = df_tl[df_tl["_ano"].isna()].reset_index(drop=True)

    LEVELS = [2.6, -2.6, 5.4, -5.4]
    _cnt   = defaultdict(int)
    card_pos = []
    for _, row in df_com.iterrows():
        ano = int(row["_ano"])
        k   = _cnt[ano]
        card_pos.append((ano, LEVELS[k % len(LEVELS)]))
        _cnt[ano] += 1

    CW, CH, CHH = 0.84, 1.85, 0.58
    X0_TL, X1_TL   = 2024.3, 2033.8
    X0_SEM, X1_SEM = 2034.6, 2037.6

    shapes, annotations = [], []
    hx, hy, ht = [], [], []

    shapes.append(dict(
        type="line", x0=X0_TL, x1=X1_TL, y0=0, y1=0,
        line=dict(color="#0D2E57", width=4),
    ))

    for ano in sorted(df_com["_ano"].unique()):
        a = int(ano)
        shapes.append(dict(
            type="circle",
            x0=a - 0.17, x1=a + 0.17, y0=-0.17, y1=0.17,
            fillcolor="#0D2E57", line=dict(color="white", width=2),
            layer="above",
        ))
        annotations.append(dict(
            x=a, y=-0.60, text=f"<b>{a}</b>",
            showarrow=False, xanchor="center", yanchor="top",
            font=dict(size=12, color="#0D2E57", family="Arial, sans-serif"),
        ))

    for i, (_, row) in enumerate(df_com.iterrows()):
        cx, cy = card_pos[i]
        cor    = _cor_tl(row.get("Rota", ""))
        x0, x1 = cx - CW / 2, cx + CW / 2
        y0, y1 = cy - CH / 2, cy + CH / 2
        yh0    = y1 - CHH

        y_link = y0 if cy > 0 else y1
        shapes.append(dict(
            type="line",
            x0=cx, x1=cx,
            y0=(0.19 if cy > 0 else -0.19), y1=y_link,
            line=dict(color=cor, width=1.5, dash="dot"),
        ))
        shapes.append(dict(
            type="rect", x0=x0, x1=x1, y0=y0, y1=yh0,
            fillcolor="white", line=dict(color="#C5D5E5", width=0.8),
            layer="above",
        ))
        shapes.append(dict(
            type="rect", x0=x0, x1=x1, y0=yh0, y1=y1,
            fillcolor=cor, line=dict(color="rgba(255,255,255,.4)", width=0.5),
            layer="above",
        ))
        annotations += [
            dict(x=cx, y=y1 - 0.09, text=f'<b>{_abrev(row.get("Projeto", ""))}</b>',
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=8, color="white", family="Arial, sans-serif")),
            dict(x=cx, y=yh0 - 0.10, text=str(row.get("Capacidade", "—")) + " m³/year",
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=7.5, color="#1A2A3A", family="Arial, sans-serif")),
            dict(x=cx, y=yh0 - 0.52, text=f'<i>{_abrev(row.get("Rota", ""), 22)}</i>',
                 showarrow=False, xanchor="center", yanchor="top",
                 font=dict(size=6.5, color="#555", family="Arial, sans-serif")),
        ]
        hx.append(cx); hy.append(cy)
        ht.append(
            f'<b>{row.get("Projeto", "")}</b><br>'
            f'Proponent: {row.get("Proponente", "")}<br>'
            f'Capacity: {row.get("Capacidade", "")} m³/year<br>'
            f'Route: {row.get("Rota", "")}<br>'
            f'Feedstock: {row.get("Feedstock", "")}<br>'
            f'Expected start: {row.get("Ano", "")}<br>'
            f'Stage: {row.get("Estagio", "")}<br>'
            f'Investment: {row.get("Investimento", "")}'
        )

    if len(df_sem) > 0:
        shapes.append(dict(
            type="line",
            x0=X0_SEM - 0.35, x1=X0_SEM - 0.35, y0=-7.8, y1=7.8,
            line=dict(color="#BBBBBB", width=1.5, dash="longdash"),
        ))
        annotations.append(dict(
            x=(X0_SEM + X1_SEM) / 2, y=7.5,
            text="<b>No defined<br>timeline</b>",
            showarrow=False, xanchor="center", yanchor="top", align="center",
            font=dict(size=11, color="#888", family="Arial, sans-serif"),
        ))
        cx_sem = (X0_SEM + X1_SEM) / 2
        for j, (_, row) in enumerate(df_sem.iterrows()):
            cy_sem = 5.5 - j * (CH + 0.40)
            cor    = _cor_tl(row.get("Rota", ""))
            xs0, xs1 = cx_sem - CW / 2, cx_sem + CW / 2
            ys0, ys1 = cy_sem - CH / 2, cy_sem + CH / 2
            yhs0     = ys1 - CHH

            shapes += [
                dict(type="rect", x0=xs0, x1=xs1, y0=yhs0, y1=ys1,
                     fillcolor=cor, line=dict(color="rgba(255,255,255,.4)", width=0.5),
                     layer="above"),
                dict(type="rect", x0=xs0, x1=xs1, y0=ys0, y1=yhs0,
                     fillcolor="#F5F5F5", line=dict(color="#C5D5E5", width=0.8),
                     layer="above"),
            ]
            annotations += [
                dict(x=cx_sem, y=ys1 - 0.09, text=f'<b>{_abrev(row.get("Projeto", ""))}</b>',
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=8, color="white", family="Arial, sans-serif")),
                dict(x=cx_sem, y=yhs0 - 0.10, text=str(row.get("Capacidade", "—")) + " m³/year",
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=7.5, color="#1A2A3A", family="Arial, sans-serif")),
                dict(x=cx_sem, y=yhs0 - 0.52, text=f'<i>{_abrev(row.get("Rota", ""), 22)}</i>',
                     showarrow=False, xanchor="center", yanchor="top",
                     font=dict(size=6.5, color="#555", family="Arial, sans-serif")),
            ]
            hx.append(cx_sem); hy.append(cy_sem)
            ht.append(
                f'<b>{row.get("Projeto", "")}</b><br>'
                f'Proponent: {row.get("Proponente", "")}<br>'
                f'Capacity: {row.get("Capacidade", "")} m³/year<br>'
                f'Route: {row.get("Rota", "")}<br>'
                f'Feedstock: {row.get("Feedstock", "")}<br>'
                f'Expected start: <b>No defined timeline</b><br>'
                f'Stage: {row.get("Estagio", "")}<br>'
                f'Investment: {row.get("Investimento", "")}'
            )

    _legenda = [
        ("Co-processing HEFA", "#1B4F8A"),
        ("Dedicated HEFA",     "#2980B9"),
        ("ATJ (Alcohol-to-Jet)", "#E67E22"),
        ("Other",              "#78909C"),
    ]
    annotations.append(dict(
        x=X0_TL + 0.05, y=-5.8, text="<b>Technology Route</b>",
        showarrow=False, xanchor="left", yanchor="top",
        font=dict(size=9.5, color="#0D2E57", family="Arial, sans-serif"),
    ))
    for k, (nome_r, cor_r) in enumerate(_legenda):
        lx = X0_TL + 0.05
        ly = -6.4 - k * 0.72
        shapes.append(dict(
            type="circle",
            x0=lx - 0.12, x1=lx + 0.12, y0=ly - 0.12, y1=ly + 0.12,
            fillcolor=cor_r, line=dict(color="white", width=1),
            layer="above",
        ))
        annotations.append(dict(
            x=lx + 0.22, y=ly, text=nome_r,
            showarrow=False, xanchor="left", yanchor="middle",
            font=dict(size=9, color="#333", family="Arial, sans-serif"),
        ))

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=hx, y=hy, mode="markers",
        marker=dict(size=44, opacity=0, symbol="square"),
        text=ht,
        hovertemplate="%{text}<extra></extra>",
        showlegend=False,
    ))
    fig.update_layout(
        shapes=shapes,
        annotations=annotations,
        title=dict(
            text=(
                "<b>Timeline — SAF Projects in Brazil</b><br>"
                '<span style="font-size:11px;color:#666">'
                "Expected operation start year · Hover over cards for details"
                "</span>"
            ),
            font=dict(size=15, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        xaxis=dict(
            range=[X0_TL - 0.5, X1_SEM + 0.3],
            showgrid=False, zeroline=False, showticklabels=False, fixedrange=True,
        ),
        yaxis=dict(
            range=[-8.2, 8.2],
            showgrid=False, zeroline=False, showticklabels=False, fixedrange=True,
        ),
        paper_bgcolor="#F4F7FB",
        plot_bgcolor="#F4F7FB",
        height=700,
        margin=dict(l=10, r=10, t=90, b=10),
        showlegend=False,
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=11, family="Arial, sans-serif"),
        ),
    )
    return fig


# ============================================================
# GRÁFICO ROSCA — PORTUGUÊS
# ============================================================

def criar_grafico_rosca_saf() -> go.Figure:
    """
    Gráfico de rosca com a capacidade acumulada por rota tecnológica no último ano
    da série (2025–2037). Labels externas com %, valor e setas conectoras.

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    PALETA = {
        "Coprocessamento": "#b2c73c",
        "HEFA":            "#3357ff",
        "ATJ":             "#107c42",
        "FT":              "#1A7F4B",
    }

    def _cor(rota):
        for k, v in PALETA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def _norm_r(r):
        r = str(r)
        if "Copro" in r: return "Coprocessamento HEFA"
        if "ATJ"   in r: return "ATJ"
        if "HEFA"  in r: return "HEFA Dedicado"
        if "FT"    in r: return "FT"
        return "Outros"

    df = _carregar_df()
    _df = df.copy()
    _df["_a"] = pd.to_numeric(_df["Ano"], errors="coerce")
    _df["_c"] = _df["Capacidade"].apply(_parse_cap)
    _df["_r"] = _df["Rota"].apply(_norm_r)
    _df = _df.dropna(subset=["_a", "_c"])
    _df["_a"] = _df["_a"].astype(int)
    _df = _df[_df["_a"].between(2025, 2037)]

    pvt = (_df.groupby(["_a", "_r"])["_c"].sum()
              .unstack(fill_value=0)
              .reindex(sorted(_df["_a"].unique()), fill_value=0))
    pvt_k   = pvt / 1_000
    pvt_cum = pvt_k.cumsum(axis=0)

    ULTIMO_ANO       = int(pvt_cum.index.max())
    DADOS_ULTIMO_ANO = pvt_cum.loc[ULTIMO_ANO]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Coprocessamento",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }
    ORDEM = ["Coprocessamento HEFA", "HEFA Dedicado", "ATJ"]
    ROTAS = [r for r in ORDEM if r in DADOS_ULTIMO_ANO.index]

    VALORES_R   = [DADOS_ULTIMO_ANO[r] for r in ROTAS]
    NAMES_R     = [LEGENDA_LABEL.get(r, r) for r in ROTAS]
    CORES_R     = [_cor(r) for r in ROTAS]
    VALOR_TOTAL = sum(VALORES_R)

    # Texto externo: nome + % + valor para cada fatia
    pcts = [v / VALOR_TOTAL * 100 for v in VALORES_R]
    custom_text = [
        f"<b>{name}</b><br><b>{pct:.1f}%</b><br>{_fmt(val)} mil m³/ano"
        for name, pct, val in zip(NAMES_R, pcts, VALORES_R)
    ]

    fig = go.Figure(data=[go.Pie(
        labels=NAMES_R,
        values=VALORES_R,
        hole=0.55,
        marker=dict(colors=CORES_R, line=dict(color="white", width=2)),
        text=custom_text,
        textinfo="text",
        textposition="outside",
        automargin=True,
        outsidetextfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        pull=[0.04] * len(ROTAS),
        hovertemplate=(
            "<b>%{label}</b><br>"
            "Capacidade: %{value:,.0f} mil m³/ano<br>"
            "Percentual: %{percent:.1f}%"
            "<extra></extra>"
        ),
    )])

    fig.update_layout(
        separators=",.",
        title=dict(
            text=(
                f"<b>Capacidade Acumulada por Rota Tecnológica — {ULTIMO_ANO}</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Distribuição da capacidade instalada acumulada (mil m³ SAF/ano)"
                "</span>"
            ),
            font=dict(size=18, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center", y=0.97,
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=680,
        margin=dict(l=60, r=60, t=130, b=80),
        showlegend=False,
        annotations=[dict(
            text=f"<b>{_fmt(VALOR_TOTAL)}</b><br>mil m³/ano",
            showarrow=False,
            font=dict(size=26, color="#333333", family="Arial, sans-serif"),
            x=0.5, y=0.5, align="center",
        )],
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=13, family="Arial, sans-serif"),
        ),
        font=dict(family="Arial, sans-serif"),
    )
    return fig


# ============================================================
# GRÁFICO ACUMULADO — PORTUGUESE
# ============================================================

def criar_grafico_acumulado_saf() -> go.Figure:
    """
    Barras empilhadas com capacidade acumulada por rota tecnológica ao longo
    dos anos de operação (2025–2037). Rótulos de total acumulado acima de cada barra.

    Fonte: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Retorna: plotly.graph_objects.Figure
    """
    PALETA = {
        "Coprocessamento": "#b2c73c",
        "HEFA":            "#3357ff",
        "ATJ":             "#107c42",
        "FT":              "#1A7F4B",
    }

    def _cor(rota):
        for k, v in PALETA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def _norm_r(r):
        r = str(r)
        if "Copro" in r: return "Coprocessamento HEFA"
        if "ATJ"   in r: return "ATJ"
        if "HEFA"  in r: return "HEFA Dedicado"
        if "FT"    in r: return "FT"
        return "Outros"

    df = _carregar_df()
    _df = df.copy()
    _df["_a"] = pd.to_numeric(_df["Ano"], errors="coerce")
    _df["_c"] = _df["Capacidade"].apply(_parse_cap)
    _df["_r"] = _df["Rota"].apply(_norm_r)
    _df = _df.dropna(subset=["_a", "_c"])
    _df["_a"] = _df["_a"].astype(int)
    _df = _df[_df["_a"].between(2025, 2037)]

    pvt = (_df.groupby(["_a", "_r"])["_c"].sum()
              .unstack(fill_value=0)
              .reindex(sorted(_df["_a"].unique()), fill_value=0))
    pvt_k   = pvt / 1_000
    pvt_cum = pvt_k.cumsum(axis=0)
    tot_cum = pvt_cum.sum(axis=1)
    Y_MAX   = tot_cum.max()

    ANOS_S = [str(a) for a in pvt_cum.index]
    ORDEM  = ["Coprocessamento HEFA", "HEFA Dedicado", "ATJ"]
    ROTAS  = [r for r in ORDEM if r in pvt_cum.columns]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Coprocessamento",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }

    fig = go.Figure()

    for rota in ROTAS:
        vals = pvt_cum[rota].values
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=LEGENDA_LABEL.get(rota, rota),
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{rota}</b><br>"
                "Ano: %{x}<br>"
                "Capacidade acumulada: <b>%{y:,.0f} mil m³/ano</b>"
                "<extra></extra>"
            ),
        ))

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_cum.values + Y_MAX * 0.035,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_cum],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ))

    fig.update_layout(
        barmode="stack",
        separators=",.",
        title=dict(
            text=(
                "<b>Projetos SAF no Brasil — Capacidade Acumulada por Rota Tecnológica</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Capacidade instalada acumulada (mil m³ SAF/ano) · barras empilhadas por ano de operação"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=560,
        margin=dict(l=80, r=30, t=95, b=120),
        legend=dict(
            title=dict(text="<b>Rota Tecnológica</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(0,0,0,0)",
            borderwidth=0,
            orientation="h",
            x=0.5, y=-0.22,
            xanchor="center",
            traceorder="normal",
        ),
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=12, family="Arial, sans-serif"),
        ),
        font=dict(family="Arial, sans-serif"),
    )
    fig.update_xaxes(
        tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
        title=dict(
            text="Ano de início de operação",
            font=dict(size=13, color="#555", family="Arial, sans-serif"),
        ),
        showgrid=False, zeroline=False,
        linecolor="#CCCCCC", linewidth=1,
    )
    fig.update_yaxes(
        title=dict(
            text="Capacidade acumulada (mil m³ SAF/ano)",
            font=dict(size=13, color="#555", family="Arial, sans-serif"),
        ),
        tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
        zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
        tickformat=",",
        rangemode="tozero",
        range=[0, Y_MAX * 1.15],
    )
    return fig


# ============================================================
# GRÁFICO ACUMULADO — ENGLISH
# ============================================================

def create_cumulative_chart_saf() -> go.Figure:
    """
    Stacked bar chart with cumulative capacity by technology route across
    operation years (2025–2037). Cumulative total labels shown above each bar.

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    PALETA = {
        "Coprocessamento": "#b2c73c",
        "HEFA":            "#3357ff",
        "ATJ":             "#107c42",
        "FT":              "#1A7F4B",
    }

    def _cor(rota):
        for k, v in PALETA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def _norm_r(r):
        r = str(r)
        if "Copro" in r: return "Coprocessamento HEFA"
        if "ATJ"   in r: return "ATJ"
        if "HEFA"  in r: return "HEFA Dedicado"
        if "FT"    in r: return "FT"
        return "Outros"

    df = _carregar_df()
    _df = df.copy()
    _df["_a"] = pd.to_numeric(_df["Ano"], errors="coerce")
    _df["_c"] = _df["Capacidade"].apply(_parse_cap)
    _df["_r"] = _df["Rota"].apply(_norm_r)
    _df = _df.dropna(subset=["_a", "_c"])
    _df["_a"] = _df["_a"].astype(int)
    _df = _df[_df["_a"].between(2025, 2037)]

    pvt = (_df.groupby(["_a", "_r"])["_c"].sum()
              .unstack(fill_value=0)
              .reindex(sorted(_df["_a"].unique()), fill_value=0))
    pvt_k   = pvt / 1_000
    pvt_cum = pvt_k.cumsum(axis=0)
    tot_cum = pvt_cum.sum(axis=1)
    Y_MAX   = tot_cum.max()

    ANOS_S = [str(a) for a in pvt_cum.index]
    ORDEM  = ["Coprocessamento HEFA", "HEFA Dedicado", "ATJ"]
    ROTAS  = [r for r in ORDEM if r in pvt_cum.columns]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Coprocessing",
        "HEFA Dedicado":        "HEFA",
        "ATJ":                  "ATJ",
    }

    fig = go.Figure()

    for rota in ROTAS:
        vals = pvt_cum[rota].values
        fig.add_trace(go.Bar(
            x=ANOS_S, y=vals,
            name=LEGENDA_LABEL.get(rota, rota),
            marker_color=_cor(rota),
            marker_line_color="white",
            marker_line_width=1.0,
            hovertemplate=(
                f"<b>{rota}</b><br>"
                "Year: %{x}<br>"
                "Cumulative capacity: <b>%{y:,.0f} thousand m³/year</b>"
                "<extra></extra>"
            ),
        ))

    fig.add_trace(go.Scatter(
        x=ANOS_S,
        y=tot_cum.values + Y_MAX * 0.035,
        mode="text",
        text=[f"<b>{_fmt(v)}</b>" if v > 0 else "" for v in tot_cum],
        textfont=dict(size=13, color="black", family="Arial, sans-serif"),
        showlegend=False, hoverinfo="skip",
    ))

    fig.update_layout(
        barmode="stack",
        separators=",.",
        title=dict(
            text=(
                "<b>SAF Projects in Brazil — Cumulative Capacity by Technological Route</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Cumulative installed capacity (thousand m³ SAF/year) · stacked bars by operation year"
                "</span>"
            ),
            font=dict(size=16, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center",
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=560,
        margin=dict(l=80, r=30, t=95, b=120),
        legend=dict(
            title=dict(text="<b>Technological Route</b>", font=dict(size=13, color="#0D2E57")),
            font=dict(size=12, color="#333", family="Arial, sans-serif"),
            bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(0,0,0,0)",
            borderwidth=0,
            orientation="h",
            x=0.5, y=-0.22,
            xanchor="center",
            traceorder="normal",
        ),
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=12, family="Arial, sans-serif"),
        ),
        font=dict(family="Arial, sans-serif"),
    )
    fig.update_xaxes(
        tickfont=dict(size=14, color="#333", family="Arial, sans-serif"),
        title=dict(
            text="Operation start year",
            font=dict(size=13, color="#555", family="Arial, sans-serif"),
        ),
        showgrid=False, zeroline=False,
        linecolor="#CCCCCC", linewidth=1,
    )
    fig.update_yaxes(
        title=dict(
            text="Cumulative capacity (thousand m³ SAF/year)",
            font=dict(size=13, color="#555", family="Arial, sans-serif"),
        ),
        tickfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
        zeroline=True, zerolinecolor="#CCCCCC", zerolinewidth=1,
        tickformat=",",
        rangemode="tozero",
        range=[0, Y_MAX * 1.15],
    )
    return fig


# ============================================================
# GRÁFICO ROSCA — ENGLISH
# ============================================================

def create_donut_chart_saf() -> go.Figure:
    """
    Donut chart with accumulated capacity by technology route in the last year
    of the series (2025–2037). Outside labels with %, value and connector arrows.

    Source: dados/SAF_EPE_10projetos_Validado_atualizado5.xlsx
    Returns: plotly.graph_objects.Figure
    """
    PALETA = {
        "Coprocessamento": "#b2c73c",
        "HEFA":            "#3357ff",
        "ATJ":             "#107c42",
        "FT":              "#1A7F4B",
    }

    def _cor(rota):
        for k, v in PALETA.items():
            if k.lower() in str(rota).lower():
                return v
        return "#607D8B"

    def _norm_r(r):
        r = str(r)
        if "Copro" in r: return "Coprocessamento HEFA"
        if "ATJ"   in r: return "ATJ"
        if "HEFA"  in r: return "HEFA Dedicado"
        if "FT"    in r: return "FT"
        return "Outros"

    df = _carregar_df()
    _df = df.copy()
    _df["_a"] = pd.to_numeric(_df["Ano"], errors="coerce")
    _df["_c"] = _df["Capacidade"].apply(_parse_cap)
    _df["_r"] = _df["Rota"].apply(_norm_r)
    _df = _df.dropna(subset=["_a", "_c"])
    _df["_a"] = _df["_a"].astype(int)
    _df = _df[_df["_a"].between(2025, 2037)]

    pvt = (_df.groupby(["_a", "_r"])["_c"].sum()
              .unstack(fill_value=0)
              .reindex(sorted(_df["_a"].unique()), fill_value=0))
    pvt_k   = pvt / 1_000
    pvt_cum = pvt_k.cumsum(axis=0)

    ULTIMO_ANO       = int(pvt_cum.index.max())
    DADOS_ULTIMO_ANO = pvt_cum.loc[ULTIMO_ANO]

    LEGENDA_LABEL = {
        "Coprocessamento HEFA": "Co-processing HEFA",
        "HEFA Dedicado":        "Dedicated HEFA",
        "ATJ":                  "ATJ",
    }
    ORDEM = ["Coprocessamento HEFA", "HEFA Dedicado", "ATJ"]
    ROTAS = [r for r in ORDEM if r in DADOS_ULTIMO_ANO.index]

    VALORES_R   = [DADOS_ULTIMO_ANO[r] for r in ROTAS]
    NAMES_R     = [LEGENDA_LABEL.get(r, r) for r in ROTAS]
    CORES_R     = [_cor(r) for r in ROTAS]
    VALOR_TOTAL = sum(VALORES_R)

    # Outside text: name + % + value for each slice
    pcts = [v / VALOR_TOTAL * 100 for v in VALORES_R]
    custom_text = [
        f"<b>{name}</b><br><b>{pct:.1f}%</b><br>{_fmt(val)} th. m³/yr"
        for name, pct, val in zip(NAMES_R, pcts, VALORES_R)
    ]

    fig = go.Figure(data=[go.Pie(
        labels=NAMES_R,
        values=VALORES_R,
        hole=0.55,
        marker=dict(colors=CORES_R, line=dict(color="white", width=2)),
        text=custom_text,
        textinfo="text",
        textposition="outside",
        automargin=True,
        outsidetextfont=dict(size=13, color="#333", family="Arial, sans-serif"),
        pull=[0.04] * len(ROTAS),
        hovertemplate=(
            "<b>%{label}</b><br>"
            "Capacity: %{value:,.0f} thousand m³/year<br>"
            "Share: %{percent:.1f}%"
            "<extra></extra>"
        ),
    )])

    fig.update_layout(
        title=dict(
            text=(
                f"<b>Accumulated Capacity by Technology Route — {ULTIMO_ANO}</b><br>"
                '<span style="font-size:12px;color:#666;">'
                "Distribution of accumulated installed capacity (thousand m³ SAF/year)"
                "</span>"
            ),
            font=dict(size=18, color="#0D2E57", family="Arial, sans-serif"),
            x=0.5, xanchor="center", y=0.97,
        ),
        paper_bgcolor="white",
        plot_bgcolor="white",
        autosize=True,
        height=680,
        margin=dict(l=60, r=60, t=130, b=80),
        showlegend=False,
        annotations=[dict(
            text=f"<b>{_fmt(VALOR_TOTAL)}</b><br>th. m³/yr",
            showarrow=False,
            font=dict(size=26, color="#333333", family="Arial, sans-serif"),
            x=0.5, y=0.5, align="center",
        )],
        hoverlabel=dict(
            bgcolor="white", bordercolor="#D0D8E4",
            font=dict(size=13, family="Arial, sans-serif"),
        ),
        font=dict(family="Arial, sans-serif"),
    )
    return fig
